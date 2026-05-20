from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.filters import Filter
import datetime
import io
import asyncio

from database.connection import AsyncSessionLocal
from database.crud import (
    get_all_doctors, get_doctor, create_doctor, deactivate_doctor,
    get_all_users, get_all_bookings, count_users, count_doctors,
    count_bookings, count_today_bookings
)
from keyboards.all_kb import admin_main_kb, doctors_remove_kb, confirm_remove_kb
from states.booking_states import AdminStates
from config import config

router = Router()


class IsAdmin(Filter):
    async def __call__(self, message: Message) -> bool:
        try:
            user_id = int(message.from_user.id)
        except (ValueError, TypeError):
            user_id = 0
        return user_id in config.admin_ids


@router.message(IsAdmin(), F.text == "📊 Statistika")
async def show_stats(message: Message):
    async with AsyncSessionLocal() as session:
        users = await count_users(session)
        doctors = await count_doctors(session)
        total_bookings = await count_bookings(session)
        today_bookings = await count_today_bookings(session)

    text = (
        "📊 <b>Klinika statistikasi</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        f"👥 Jami mijozlar: <b>{users}</b>\n"
        f"👨‍⚕️ Faol shifokorlar: <b>{doctors}</b>\n"
        f"📋 Jami navbatlar: <b>{total_bookings}</b>\n"
        f"📅 Bugungi navbatlar: <b>{today_bookings}</b>\n\n"
        f"📆 <i>Sana: {datetime.date.today().strftime('%d.%m.%Y')}</i>"
    )
    await message.answer(text, parse_mode="HTML")


@router.message(IsAdmin(), F.text == "📋 Barcha buyurtmalar")
async def all_orders(message: Message):
    async with AsyncSessionLocal() as session:
        bookings = await get_all_bookings(session)

    if not bookings:
        await message.answer("📭 Hali birorta navbat yo'q.")
        return

    text = "📋 <b>So'nggi 20 ta navbat:</b>\n━━━━━━━━━━━━━━━━━━\n\n"
    for b in bookings[:20]:
        text += (
            f"#{b.id} | {b.booking_date.strftime('%d.%m')} {b.booking_time.strftime('%H:%M')}\n"
            f"👤 {b.user.full_name} → 👨‍⚕️ {b.doctor.full_name}\n"
            f"🦷 {b.service} | {b.status.value}\n\n"
        )
    await message.answer(text, parse_mode="HTML")


@router.message(IsAdmin(), F.text == "📥 Excel yuklash")
async def export_excel(message: Message):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        await message.answer("❌ openpyxl o'rnatilmagan.")
        return

    async with AsyncSessionLocal() as session:
        users = await get_all_users(session)
        bookings = await get_all_bookings(session)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Mijozlar"
    headers1 = ["ID", "Ism", "Telefon", "Username", "Ro'yxatdan o'tgan"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
    for u in users:
        ws1.append([
            u.id, u.full_name, u.phone,
            f"@{u.username}" if u.username else "",
            u.created_at.strftime("%d.%m.%Y") if u.created_at else ""
        ])

    ws2 = wb.create_sheet("Navbatlar")
    headers2 = ["ID", "Mijoz", "Telefon", "Shifokor", "Sana", "Vaqt", "Xizmat", "Holat"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
    for b in bookings:
        ws2.append([
            b.id, b.user.full_name, b.user.phone,
            b.doctor.full_name,
            b.booking_date.strftime("%d.%m.%Y"),
            b.booking_time.strftime("%H:%M"),
            b.service, b.status.value
        ])

    for ws in [ws1, ws2]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"dental_bot_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"📥 <b>Ma'lumotlar tayyor!</b>\n{datetime.date.today().strftime('%d.%m.%Y')} holati",
        parse_mode="HTML"
    )


@router.message(IsAdmin(), F.text == "💌 Eslatma yuborish")
async def broadcast_start(message: Message, state: FSMContext):
    await message.answer(
        "✍️ <b>Xabar yozing</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        "Barcha mijozlarga yuboriladigan xabarni kiriting.\n"
        "<i>Maslahat: qisqa, aniq va foydali xabar ko'proq o'qiladi!</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.broadcasting)


@router.message(AdminStates.broadcasting, F.text)
async def do_broadcast(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    async with AsyncSessionLocal() as session:
        users = await get_all_users(session)

    sent = 0
    failed = 0
    for user in users:
        try:
            await bot.send_message(user.telegram_id, message.text, disable_notification=True)
            sent += 1
            if sent % 20 == 0:
                await asyncio.sleep(0.5)
        except Exception:
            failed += 1

    await message.answer(
        f"✅ <b>Xabar yuborildi!</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        f"📤 Muvaffaqiyatli: <b>{sent}</b>\n"
        f"❌ Yuborilmadi: <b>{failed}</b>",
        parse_mode="HTML",
        reply_markup=admin_main_kb()
    )


@router.message(IsAdmin(), F.text == "💬 Xabar yozish")
async def message_user_start(message: Message, state: FSMContext):
    await message.answer(
        "👤 <b>Foydalanuvchi Telegram ID sini kiriting:</b>\n\n"
        "<i>ID ni bilmasa, foydalanuvchi @userinfobot orqali bilib olishi mumkin</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.messaging_user)


# BUG FIX: 2 bosqichli messaging_user — birinchi ID, keyin xabar
# Avvalgi kodda ikkinchi handler state filtrisiz bo'lgani uchun
# birinchi xabar ham "xabar" sifatida qabul qilinardi.
# Yechim: alohida state (messaging_user_text) ishlatish.

@router.message(AdminStates.messaging_user, F.text)
async def message_user_get_id(message: Message, state: FSMContext):
    try:
        user_id = int(message.text.strip())
    except (ValueError, TypeError):
        await message.answer("❌ Noto'g'ri ID. Faqat raqam kiriting:")
        return

    await state.update_data(target_user=user_id)
    await message.answer(
        f"✍️ <b>{user_id}</b> ga yuboriladigan xabarni kiriting:"
    )
    await state.set_state(AdminStates.messaging_user_text)


@router.message(AdminStates.messaging_user_text, F.text)
async def message_user_send(message: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    target_id = data.get("target_user")
    if not target_id:
        await message.answer("Xatolik yuz berdi. Qayta urinib ko'ring.")
        await state.clear()
        return

    try:
        await bot.send_message(target_id, message.text)
        await message.answer(
            f"✅ <b>Xabar yuborildi!</b>\n\n"
            f"Qabul qiluvchi: <code>{target_id}</code>",
            parse_mode="HTML",
            reply_markup=admin_main_kb()
        )
    except Exception as e:
        await message.answer(
            f"❌ <b>Xabar yuborilmadi!</b>\n\n<code>{e}</code>",
            parse_mode="HTML",
            reply_markup=admin_main_kb()
        )
    await state.clear()


@router.message(IsAdmin(), F.text == "➕ Shifokor qo'shish")
async def add_doctor_start(message: Message, state: FSMContext):
    await message.answer(
        "👨‍⚕️ <b>Yangi shifokor qo'shish</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        "Shifokorning to'liq ismini kiriting\n"
        "<i>(Familiya Ism Otasining ismi)</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.adding_doctor_name)


@router.message(AdminStates.adding_doctor_name, F.text)
async def got_doctor_name(message: Message, state: FSMContext):
    await state.update_data(doctor_name=message.text.strip())
    await message.answer("🩺 Mutaxassisligini kiriting:\n<i>(masalan: Terapevt, Xirurg, Ortodont)</i>", parse_mode="HTML")
    await state.set_state(AdminStates.adding_doctor_specialty)


@router.message(AdminStates.adding_doctor_specialty, F.text)
async def got_doctor_specialty(message: Message, state: FSMContext):
    await state.update_data(doctor_specialty=message.text.strip())
    await message.answer("📱 Telefon raqamini kiriting:\n<i>(yoki «o'tkazish» yozing)</i>", parse_mode="HTML")
    await state.set_state(AdminStates.adding_doctor_phone)


@router.message(AdminStates.adding_doctor_phone, F.text)
async def got_doctor_phone(message: Message, state: FSMContext):
    phone = None if message.text.strip().lower() == "o'tkazish" else message.text.strip()
    await state.update_data(doctor_phone=phone)
    await message.answer(
        "🆔 Shifokorning Telegram ID sini kiriting:\n"
        "<i>(bilmasangiz «0» yozing — keyinroq qo'shsa bo'ladi)</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.adding_doctor_tg)


@router.message(AdminStates.adding_doctor_tg, F.text)
async def got_doctor_tg(message: Message, state: FSMContext):
    data = await state.get_data()
    tg_id_raw = message.text.strip().replace(" ", "")
    try:
        tg_id = int(tg_id_raw) if tg_id_raw.isdigit() and int(tg_id_raw) != 0 else None
    except (ValueError, TypeError):
        tg_id = None

    async with AsyncSessionLocal() as session:
        doc = await create_doctor(
            session,
            full_name=data["doctor_name"],
            specialty=data["doctor_specialty"],
            phone=data.get("doctor_phone"),
            telegram_id=tg_id
        )

    await state.clear()
    await message.answer(
        f"✅ <b>Shifokor muvaffaqiyatli qo'shildi!</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        f"👨‍⚕️ <b>{doc.full_name}</b>\n"
        f"🩺 {doc.specialty}\n"
        f"📱 {doc.phone or '—'}\n"
        f"🆔 {tg_id or '—'}\n\n"
        f"Endi mijozlar ushbu shifokorga navbat ola oladi!",
        parse_mode="HTML",
        reply_markup=admin_main_kb()
    )


@router.message(IsAdmin(), F.text == "🗑️ Shifokor o'chirish")
async def remove_doctor_start(message: Message):
    async with AsyncSessionLocal() as session:
        doctors = await get_all_doctors(session)
    if not doctors:
        await message.answer("📭 Hozircha shifokorlar yo'q.")
        return
    await message.answer(
        "🗑️ <b>Qaysi shifokorni o'chirish?</b>\n━━━━━━━━━━━━━━━━━━",
        parse_mode="HTML",
        reply_markup=doctors_remove_kb(doctors)
    )


@router.callback_query(F.data.startswith("remove_doctor:"))
async def confirm_remove(callback: CallbackQuery):
    doctor_id = int(callback.data.split(":")[1])
    async with AsyncSessionLocal() as session:
        doc = await get_doctor(session, doctor_id)
    await callback.answer()
    await callback.message.edit_text(
        f"⚠️ <b>Diqqat!</b>\n\n"
        f"<b>{doc.full_name}</b> ni tizimdan o'chirmoqchimisiz?\n"
        f"Bu amalni qaytarib bo'lmaydi!",
        parse_mode="HTML",
        reply_markup=confirm_remove_kb(doctor_id)
    )


@router.callback_query(F.data.startswith("confirm_remove:"))
async def do_remove_doctor(callback: CallbackQuery):
    doctor_id = int(callback.data.split(":")[1])
    async with AsyncSessionLocal() as session:
        doc = await get_doctor(session, doctor_id)
        await deactivate_doctor(session, doctor_id)
    await callback.answer("O'chirildi!")
    await callback.message.edit_text(
        f"✅ <b>{doc.full_name}</b> tizimdan o'chirildi.",
        parse_mode="HTML"
    )


@router.callback_query(F.data == "cancel_admin")
async def cancel_admin_action(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.answer()
    await callback.message.edit_text("❌ Bekor qilindi.")


# BUG FIX: admin_fallback faqat state yo'q bo'lganda ishlasin
# Avval: IsAdmin filtri yo'q edi — har qanday xabar ushbu handlerga tushar edi
@router.message(IsAdmin())
async def admin_fallback(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        await message.answer(
            "🔐 <b>Admin panel</b>",
            parse_mode="HTML",
            reply_markup=admin_main_kb()
        )
